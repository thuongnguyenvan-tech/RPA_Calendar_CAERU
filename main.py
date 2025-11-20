from PyQt5 import QtCore, QtGui, QtWidgets
import sys
import os
from PyQt5.QtWidgets import QApplication, QWidget, QComboBox, QPushButton, QFileDialog, QVBoxLayout
import pandas as pd
from playwright.sync_api import sync_playwright
import os
import re
import json
import sys
import pandas as pd
from collections import defaultdict
from itertools import islice
from datetime import datetime
from prompt import CHECK_PROMPT
from bs4 import BeautifulSoup
import ctypes
from openpyxl import Workbook, load_workbook
import os
import time
from datetime import datetime

class Ui_MainWindow(object):
    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(800, 250)
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.formLayoutWidget = QtWidgets.QWidget(self.centralwidget)
        self.formLayoutWidget.setGeometry(QtCore.QRect(6, 1, 791, 351))
        self.formLayoutWidget.setObjectName("formLayoutWidget")
        self.formLayout = QtWidgets.QFormLayout(self.formLayoutWidget)
        self.formLayout.setContentsMargins(0, 0, 0, 0)
        self.formLayout.setObjectName("formLayout")
        self.horizontalLayout = QtWidgets.QHBoxLayout()
        self.horizontalLayout.setObjectName("horizontalLayout")
        self.label = QtWidgets.QLabel(self.formLayoutWidget)
        self.label.setObjectName("label")
        self.horizontalLayout.addWidget(self.label)
        self.verticalLayout = QtWidgets.QVBoxLayout()
        self.verticalLayout.setObjectName("verticalLayout")
        self.pushButton = QtWidgets.QPushButton(self.formLayoutWidget)
        self.pushButton.setObjectName("pushButton")
        self.verticalLayout.addWidget(self.pushButton)
        self.horizontalLayout.addLayout(self.verticalLayout)
        self.formLayout.setLayout(0, QtWidgets.QFormLayout.FieldRole, self.horizontalLayout)
        self.horizontalLayout_2 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_2.setObjectName("horizontalLayout_2")
        self.label_4 = QtWidgets.QLabel(self.formLayoutWidget)
        self.label_4.setObjectName("label_4")
        self.horizontalLayout_2.addWidget(self.label_4)
        self.label_3 = QtWidgets.QLabel(self.formLayoutWidget)
        self.label_3.setObjectName("label_3")
        self.horizontalLayout_2.addWidget(self.label_3)
        self.formLayout.setLayout(2, QtWidgets.QFormLayout.FieldRole, self.horizontalLayout_2)
        self.verticalLayout_2 = QtWidgets.QVBoxLayout()
        self.verticalLayout_2.setObjectName("verticalLayout_2")
        self.label_6 = QtWidgets.QLabel(self.formLayoutWidget)
        self.label_6.setObjectName("label_6")
        self.verticalLayout_2.addWidget(self.label_6)
        self.label_7 = QtWidgets.QLabel(self.formLayoutWidget)
        self.label_7.setObjectName("label_7")
        self.verticalLayout_2.addWidget(self.label_7)
        self.formLayout.setLayout(3, QtWidgets.QFormLayout.FieldRole, self.verticalLayout_2)
        self.verticalLayout_3 = QtWidgets.QVBoxLayout()
        self.verticalLayout_3.setObjectName("verticalLayout_3")
        self.label_8 = QtWidgets.QLabel(self.formLayoutWidget)
        self.label_8.setObjectName("label_8")
        self.verticalLayout_3.addWidget(self.label_8)
        self.label_9 = QtWidgets.QLabel(self.formLayoutWidget)
        self.label_9.setObjectName("label_9")
        self.verticalLayout_3.addWidget(self.label_9)
        self.horizontalLayout_3 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_3.setObjectName("horizontalLayout_3")
        self.pushButton_2 = QtWidgets.QPushButton(self.formLayoutWidget)
        self.pushButton_2.setObjectName("pushButton_2")
        self.horizontalLayout_3.addWidget(self.pushButton_2)
        self.pushButton_3 = QtWidgets.QPushButton(self.formLayoutWidget)
        self.pushButton_3.setObjectName("pushButton_3")
        self.horizontalLayout_3.addWidget(self.pushButton_3)
        self.pushButton_4 = QtWidgets.QPushButton(self.formLayoutWidget)
        self.pushButton_4.setObjectName("pushButton_4")
        self.horizontalLayout_3.addWidget(self.pushButton_4)
        self.verticalLayout_3.addLayout(self.horizontalLayout_3)
        self.formLayout.setLayout(4, QtWidgets.QFormLayout.FieldRole, self.verticalLayout_3)
        self.horizontalLayout_4 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_4.setObjectName("horizontalLayout_4")
        self.label_2 = QtWidgets.QLabel(self.formLayoutWidget)
        self.label_2.setObjectName("label_2")
        self.horizontalLayout_4.addWidget(self.label_2)
        self.comboBox = QtWidgets.QComboBox(self.formLayoutWidget)
        self.comboBox.setObjectName("comboBox")
        self.comboBox.addItem("")
        self.comboBox.addItem("")
        self.horizontalLayout_4.addWidget(self.comboBox)
        self.formLayout.setLayout(1, QtWidgets.QFormLayout.FieldRole, self.horizontalLayout_4)
        MainWindow.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(MainWindow)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 800, 22))
        self.menubar.setObjectName("menubar")
        MainWindow.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)
        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "MainWindow"))
        self.label.setText(_translate("MainWindow", "File name"))
        self.pushButton.setText(_translate("MainWindow", "Upload File"))
        self.label_4.setText(_translate("MainWindow", "State:"))
        self.label_3.setText(_translate("MainWindow", "Waiting"))
        self.label_6.setText(_translate("MainWindow", "Current Step:"))
        self.label_7.setText(_translate("MainWindow", "..."))
        self.label_8.setText(_translate("MainWindow", "Error:"))
        self.label_9.setText(_translate("MainWindow", "..."))
        self.pushButton_2.setText(_translate("MainWindow", "Start"))
        self.pushButton_3.setText(_translate("MainWindow", "Stop"))
        self.pushButton_4.setText(_translate("MainWindow", "Exit"))
        self.label_2.setText(_translate("MainWindow", "Headless:"))
        self.comboBox.setItemText(0, _translate("MainWindow", "True"))
        self.comboBox.setItemText(1, _translate("MainWindow", "False"))
        self.pushButton.clicked.connect(lambda: self.upload_file(MainWindow))
        self.pushButton_4.clicked.connect(QtWidgets.QApplication.quit)
        self.pushButton_2.clicked.connect(self.click_start)
        self.pushButton_3.clicked.connect(self.click_stop)
        
    def upload_file(self, parent):
        file_filter = "Excel Files (*.xlsx *.csv)"
        self.file_path, _ = QFileDialog.getOpenFileName(
            parent=parent,
            caption="Select a data file",
            directory=os.getcwd(),
            filter=file_filter,
            initialFilter="Excel Files (*.xlsx *.csv)"
        )
        if self.file_path:
            self.label.setText(self.file_path)
    def enable_buttons(self):
        self.pushButton_2.setEnabled(True)
        self.pushButton_4.setEnabled(True)
        self.pushButton.setEnabled(True)

    def click_start(self):
        self.label_3.setText("Processing")
        self.label_3.adjustSize()
        self.pushButton_4.setEnabled(False)
        self.pushButton_2.setEnabled(False)
        self.pushButton.setEnabled(False)
        self.headless = self.comboBox.currentText()
        self.worker = Playwright(self.file_path, self.headless)
        self.worker.error.connect(self.on_error)
        self.worker.current_step.connect(self.on_current_step)
        self.worker.status.connect(self.on_status)
        self.worker.button.connect(self.enable_buttons)
        self.worker.start()

    def click_stop(self):
        self.worker.stop()
        self.label_3.setText("Stopping...")
        self.label_3.adjustSize()
        QtCore.QTimer.singleShot(3000, self.reset)

    def reset(self):
        self.label_3.setText("Waiting")
        self.label_3.adjustSize()

    def on_error(self, text):
        if text == "__CLEAR__":
            self.label_9.clear()
            return
        self.label_9.clear()
        self.label_9.setText(text)
    def on_current_step(self, text):
        if text == "__CLEAR__":
            self.label_7.clear()
            return
        self.label_7.clear()
        self.label_7.setText(text)
    def on_status(self, text):
        self.label_3.clear()
        self.label_3.setText(text)

class Playwright(QtCore.QThread):
    error = QtCore.pyqtSignal(str) 
    current_step = QtCore.pyqtSignal(str) 
    status = QtCore.pyqtSignal(str) 
    button = QtCore.pyqtSignal() 

    def __init__(self, file_path, headless, parent=None):
        super().__init__(parent)
        self.file_path = file_path
        self.headless = headless
        self._is_running = True


    def group_dates_by_year_month(self, date_list):
        result = defaultdict(lambda: defaultdict(list))
        for date_str in date_list:
            year, month, _ = date_str.split('-')
            month = int(month)
            year = int(year)
            result[year][month].append(date_str)
        result = {year: dict(months) for year, months in result.items()}
        return result

    def get_holidays_weekends(self, df):
        result = {}
        for category in df['種別'].unique():
            result[category] = df.loc[df['種別'] == category, '日付'].tolist()
        return result

    def circled_number_to_int(self, c):
        return ord(c) - ord('①') + 1

    def parse_pattern_sheet_numeric(self, df):
        result = {}
        num_cols = df.shape[1]
        for col in range(0, num_cols, 2):
            pattern_name = df.iloc[0, col]
            if pd.isna(pattern_name):
                continue
            pattern_number = self.circled_number_to_int(pattern_name[0])
            pattern_dict = {"出勤": [], "休日": []}
            day_type_row = df.iloc[1, col]
            dates = df.iloc[2:, col].dropna()
            dates = [d.strftime("%Y-%m-%d") if hasattr(d, "strftime") else str(d) for d in dates]
            pattern_dict[day_type_row].extend(dates)
            if col + 1 < num_cols:
                adj_day_type = df.iloc[1, col + 1]
                if pd.notna(adj_day_type):
                    adj_dates = df.iloc[2:, col + 1].dropna()
                    adj_dates = [d.strftime("%Y-%m-%d") if hasattr(d, "strftime") else str(d) for d in adj_dates]
                    pattern_dict[adj_day_type].extend(adj_dates)
            result[pattern_number] = pattern_dict
        return result

    def extract_td_tags(self, html):
        soup = BeautifulSoup(html, 'html.parser')
        td_tags = soup.find_all('td')
        td_strings = [str(td) for td in td_tags if td.get_text(strip=True)]
        return td_strings

    def check_calendar_days(self, calendar_html, schedule, current_year, month):
        results = []
        for td_html in calendar_html:
            soup = BeautifulSoup(td_html, 'html.parser')
            td = soup.find('td')
            day_text = td.text.strip()
            if not day_text.isdigit():
                results.append(False)
                continue
            date_str = f"{current_year}-{month:02d}-{int(day_text):02d}"
            classes = td.get('class', [])
            if "pink_holiday" in classes:
                results.append(date_str in schedule[current_year][month]['red_days'])
            elif "blue_holiday" in classes:
                results.append(date_str in schedule[current_year][month]['blue_days'])
            elif "pointable" in classes:
                results.append(date_str in schedule[current_year][month]['black_days'])
            else:
                results.append(False)
        return results

    def find_day_category(self, day_str, schedule):
        for key, days in schedule.items():
            if day_str in days:
                return key
        return None

    def get_click_count(self, td_html, target_status):
        if "pink_holiday" in td_html:
            current_status = "red_days"
        elif "blue_holiday" in td_html:
            current_status = "blue_days"
        else:
            current_status = "black_days"

        states = ["red_days", "blue_days", "black_days"]

        current_idx = states.index(current_status)
        target_idx = states.index(target_status)

        clicks = (target_idx - current_idx) % len(states)
        return clicks

    def extract_info_from_info_sheet(self, df):
        df = df.copy()
        df.columns = [c.strip() for c in df.columns]
        info_map = df.set_index('name')['value'].to_dict()

        def get(key):
            v = info_map.get(key)
            if pd.isna(v):
                return None
            return str(v).strip()
        
        url = get('URL')
        admin_id = get('管理者ID')
        password = get('password')

        raw_year = get('year')
        years = None
        if raw_year:
            parts = [p.strip() for p in raw_year.replace('，', ',').split(',')]
            parts = [p for p in parts if p]
            parsed = []
            for p in parts:
                try:
                    parsed.append(int(p))
                except ValueError:
                    parsed.append(p)
            years = parsed

        return {
            'URL': url,
            '管理者ID': admin_id,
            'password': password,
            'year': years
        }

    def extract_weekend_days(self, tr_html_list, year_month):
        result = {"blue_days": [], "red_days": [], "black_days": []}
        table = [] 
        for tr_html in tr_html_list:
            soup = BeautifulSoup(tr_html, "html.parser")
            tds = soup.find_all("td")
            row = []
            for td in tds:
                text = td.get_text(strip=True)
                row.append(text)
            table.append(row)
        for weekday in table[1:]:
            for index, value in enumerate(weekday):
                if value != '' and index == 0:
                    result["red_days"].append(f"{year_month}{int(weekday[index]):02d}")
                if value != '' and index == 6:
                    result["blue_days"].append(f"{year_month}{int(weekday[index]):02d}")
                if value != '' and index != 0 and index != 6:
                    result["black_days"].append(f"{year_month}{int(weekday[index]):02d}")
        return result

    def update_day_patterns(self, A, B):
        holidays = set(A.get("red_days", []))
        leave_days = set(A.get("blue_days", []))
        work_days = set(A.get("black_days", []))
        sundays = set(B.get("red_days", []))
        saturdays = set(B.get("blue_days", []))
        normal_days = set(B.get("black_days", []))
        A["red_days"] = sorted((sundays | holidays) - work_days)
        A["blue_days"] = sorted((saturdays | leave_days) - work_days - holidays)
        A["black_days"] = sorted((normal_days | work_days)- leave_days - holidays) 
        return A

    def generate_year_click_code(self, target_year, current_year):
        diff = target_year - current_year
        if diff == 0:
            return 0
        if diff > 0:
            clicks = "\n".join(["button.nth(1).click()" for _ in range(diff)])
        elif diff < 0:
            clicks = "\n".join(["button.first.click()" for _ in range(abs(diff))])
        return clicks

    def create_sheet_result_output(self, file_path,years):
        wb = load_workbook(file_path)
        if "result" not in wb.sheetnames:
            wb.create_sheet("result")
        else:
            del wb["result"]
            wb.create_sheet("result")
        ws = wb["result"]
        ws.append(["勤務地名", "勤務地ID", "パターン選択"] + years)
        # Lưu file
        wb.save(file_path)

    def add_row_sheet_result_output(self, file_path, rows):
        wb = load_workbook(file_path)
        ws = wb["result"]
        ws.append(rows)
        wb.save(file_path)

    def add_block_sheet_result_output(self, file_path, row, column, value):
        wb = load_workbook(file_path)
        ws = wb["result"]
        ws.cell(row=row, column=column, value=value)
        wb.save(file_path)

    def validate_dates(self, date_list):
        invalid = []
        for d in date_list:
            try:
                datetime.strptime(d, "%Y-%m-%d")
            except ValueError:
                invalid.append(d)
        if len(invalid) != 0:
            self.reset(f"# Error: Date format in sheet 2 named 法定休日: " + "".join(invalid))
            return False
        else:
            return True
    def reset(self, text):
        self.error.emit(text)
        self.current_step.emit("__CLEAR__")
        self.status.emit(f"Waiting")

    def stop(self):
        self._is_running = False
        self.button.emit() 
        self.error.emit("__CLEAR__") 
        self.current_step.emit("__CLEAR__")

    def run(self):
        self.error.emit("__CLEAR__")

        self.current_step.emit("Preprocessing data")
        try:
            sheet1 = pd.read_excel(self.file_path, sheet_name='情報')
        except Exception as e:
            self.reset(f"Input excel file don't have a sheet named 情報. Please check again and upload again.")
            return
        try:
            sheet2 = pd.read_excel(self.file_path, sheet_name='法定休日')
        except Exception as e:
            self.reset(f"Input excel file don't have a sheet named 法定休日. Please check again and upload again.")
            return
        try:
            sheet3 = pd.read_excel(self.file_path, sheet_name='一般休日')
        except Exception as e:
            self.reset(f"Input excel file don't have a sheet named 一般休日. Please check again and upload again.")
            return
        try:
            sheet4 = pd.read_excel(self.file_path, sheet_name='パターン内容', parse_dates=True)
        except Exception as e:
            self.reset(f"Input excel file don't have a sheet named パターン内容. Please check again and upload again.")
            return

        self.current_step.emit(f"Success in reading input excel file.")

        #Sheet 1
        information = self.extract_info_from_info_sheet(sheet1)

        #Check parameter year trong sheet 1
        for year in information['year']:
            if not (2000 <= year <= 2100):
                self.reset("# Error: year out of range 2000-2100")
                return
        self.create_sheet_result_output(file_path = self.file_path, years = information['year'])

        #Sheet 2 
        result_data = self.validate_dates(self.get_holidays_weekends(sheet2)["祝日"])
        if result_data == False:
            return
        leave_schedule = self.group_dates_by_year_month(self.get_holidays_weekends(sheet2)["祝日"])

        #Sheet 3
        butans = sheet3.iloc[:, 2]
        if len(butans) == 0:
            self.reset(f"# Error: Date format in sheet 3 named 一般休日. Column パターン選択 do not have any value. Check again.")
            return
        invalid = []
        valid = []
        for butan in butans:
            butan = str(butan)
            if '\u2460' <= butan <= '\u2473':
                invalid.append(butan)
            else:
                valid.append(butan)

        if len(valid)!=0 and len(invalid)!=0:
            self.reset(f"# Error: Date format in sheet 3 named 一般休日. Column パターン選択 do not have the same type (circled numbers or numbers). Check again.")
            return
        
        if len(invalid) == len(butans):
            sheet3['パターン選択'] = sheet3['パターン選択'].apply(self.circled_number_to_int)
        elif len(valid) == len(butans):
            pass
        else:
            self.reset(f"# Error: Date format in sheet 3 named 一般休日. Check again.")
            return

        invalid_values = [v for v in sheet3.iloc[:, 2] if not str(v).isdigit()]

        if len(invalid_values) != 0:
            self.reset(f"# Error: Format Error in sheet 3 named 一般休日. Column パターン選択 have value that is not digital. Check again.")
            return
        areas = sheet3.set_index('勤務地ID')['パターン選択'].to_dict()
        kinmuchi = sheet3.set_index('勤務地ID')['勤務地名'].to_dict()

        #Sheet 4
        butan_list = self.parse_pattern_sheet_numeric(sheet4)

        #Final area list
        final_area_list = {k: butan_list[v] for k, v in areas.items()}

        # final_area_list = dict(islice(final_area_list.items(), 5))

        master_schedule = {}

        for area_id, pattern_data in final_area_list.items():
            master_schedule[area_id] = defaultdict(lambda: defaultdict(dict))
            # Check date format
            for list_dates in pattern_data.values():
                result_data = self.validate_dates(list_dates)
                if result_data == False:
                    return
            for year in leave_schedule.keys():
                for month in range(1, 13):
                    leave_days = [
                            day for day in pattern_data.get("休日", [])
                            if datetime.strptime(day, "%Y-%m-%d").year == year and
                            datetime.strptime(day, "%Y-%m-%d").month == month
                        ]
                    work_days = [
                            day for day in pattern_data.get("出勤", [])
                            if datetime.strptime(day, "%Y-%m-%d").year == year and
                            datetime.strptime(day, "%Y-%m-%d").month == month
                        ]
                    
                    holidays = leave_schedule.get(year, {}).get(month, [])

                    red_days = sorted(set(holidays))

                    blue_days = sorted(set(leave_days))

                    black_days = sorted(set(work_days))

                    master_schedule[area_id][year][month] = {
                        "red_days": red_days,
                        "blue_days": blue_days,
                        "black_days": black_days
                    }

        master_schedule = {k: dict(v) for k, v in master_schedule.items()}

        self.current_step.emit("Login")
        with sync_playwright() as playwright:
            headless_bool = (self.headless == "True")
            browser = playwright.chromium.launch(headless=headless_bool, slow_mo=500,args=["--kiosk"])
            page = browser.new_page(viewport={"width": 1920, "height": 1080})
            page.goto(information['URL'], wait_until='domcontentloaded')
            try:
                page.fill('input[name="presentation_id"]', information['管理者ID'])
            except Exception as e:
                self.reset(f"Can not login. Please check again the URL in sheet 1 named 情報.")
                return
            
            page.fill('input[name="password"]', information['password'])
            page.click('button:has-text("ログイン")')
            page.wait_for_load_state("domcontentloaded")
            if page.locator("main#home_page").is_visible():
                pass
            else:
                self.reset(f"Can not login. Please check again the 管理者ID and password in sheet 1 named 情報.")
                return
            page.goto(information['URL'] + "calendar", wait_until="domcontentloaded")

            first_loop = True
            row = 2

            #TODO: Thêm bước xác nhận nhập account có đúng không ở đây


            for area_ID, schedule in master_schedule.items():
                # area_ID = "S02000"
                column = 4
                self.add_row_sheet_result_output(self.file_path, [kinmuchi.get(area_ID), area_ID, areas.get(area_ID)])
                if not first_loop:
                    page.locator('a.modal-open.btn_gray', has_text="変更").click()
                else:
                    first_loop = False
                input_locator = page.locator('div.search_setting:has(span:text-is("勤務地ID")) input')
                input_locator.fill(area_ID)
                input_locator.press("Enter")
                try:
                    second_button = page.locator('a.ss_size.s_height.btn_gray:text-is("選択")').nth(1)
                except Exception as e:
                    self.add_block_sheet_result_output(file_path =self.file_path, row = row, column = column, value = f"Không có mã kinmuchi {area_ID}")
                    first_loop = True
                    row += 1
                    continue
                second_button.click()
                
                for target_year in information['year']:
                    # target_year = 2026
                    year_text = page.locator('section.ll_font').inner_text()
                    current_year = int(year_text.replace("年", "").strip())
                    button = page.locator('section.ico_position img.ico_ico_arrow')

                    clicks = self.generate_year_click_code(target_year = target_year, current_year = current_year)
                    if clicks != 0:
                        exec(clicks)
                    page.wait_for_load_state("domcontentloaded")
                    time.sleep(2)
                    error_months = []

                    for month in range(1,13):
                        error_found = False

                        if self._is_running == False:
                            return
                        self.current_step.emit(f"Checking 勤務地名{area_ID} - year {target_year} - month {month}")
                        # month = 1
                        current_calendar = page.locator('section.caeru_calendar_wrapper').filter(has=page.locator(f'span:text-is("{month}月")'))
                        tr_elements = current_calendar.locator('table tr')

                        # Lấy nội dung HTML từng <tr>
                        tr_html_list = [tr_elements.nth(i).inner_html() for i in range(1,tr_elements.count())]
                        year_month = f"{target_year}-{month:02d}-"
                        weekends = (self.extract_weekend_days(tr_html_list,year_month))
                        schedule[target_year][month] = self.update_day_patterns(schedule[target_year][month], weekends)

                        # with open("master_schedule.txt", "w", encoding="utf-8") as f:
                        #     json.dump(schedule[target_year][month], f, ensure_ascii=False, indent=4)

                        calendar_html = current_calendar.inner_html()
                        td_list = self.extract_td_tags(calendar_html)[7:]
                        
                        not_done_before = "pink_holiday" not in "".join(td_list) and "pink_holiday" not in "".join(td_list)

                        if not_done_before:
                            for type_day, list_type_days in schedule[target_year][month].items():
                                try: 
                                    for leaves in list_type_days:
                                        day = leaves[-2:]
                                        day_cell = current_calendar.locator(f'td.pointable:text-is("{int(day)}")')
                                        if type_day == "red_days":
                                            day_cell.click()
                                        elif type_day == "blue_days":
                                            day_cell.dblclick()
                                except Exception as e:
                                    print(f"Không có {type_day} ở tháng thứ {month}.Error: {e}")
                                    self.add_block_sheet_result_output(file_path =self.file_path, row = row, column = column, value = f"Error at month {month}, so we still not save it. Please check again. ")
                                    row += 1
                                    error_found = True
                                    break
                        if error_found:
                            continue 
                        calendar_html = current_calendar.inner_html()
                        td_list = self.extract_td_tags(calendar_html)[7:]
                        result = self.check_calendar_days(td_list, schedule, target_year, month)

                        if False in result:
                            print(f"Tháng {month}: Error!")
                            for index, day in enumerate(result):
                                if day == False:
                                    print(f"Lỗi ở ngày {index+1}")
                                    current_error_day_status = td_list[index]
                                    date_str = f"{target_year}-{month:02d}-{int(index+1):02d}"
                                    target = self.find_day_category(date_str, schedule[target_year][month])
                                    day_cell = current_calendar.locator(f'td.pointable:text-is("{int(index+1)}")')
                                    clicks = self.get_click_count(current_error_day_status, target)
                                    if clicks == 1:
                                        day_cell.click()
                                    elif clicks == 2:
                                        day_cell.dblclick()
                                    print(f"Đã sửa ngày {index+1} thành {target}")
                            
                        else:
                            print(f"Tháng {month}: OK!")

                        if not not_done_before and (False in result):
                            save_button = current_calendar.locator('a.btn_greeen:has-text("保存")')
                            save_button.click()
                            page.wait_for_load_state("domcontentloaded")
                    
                    if len(error_months) != 0:
                        self.add_block_sheet_result_output(file_path =self.file_path, row = row, column = column, value = "Error with " + ", ".join(map(str, error_months)))
                    else:
                        self.add_block_sheet_result_output(file_path =self.file_path, row = row, column = column, value = "Done 12 months without error !")
                        
                    column += 1

                row += 1

if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = Ui_MainWindow()
    ui.setupUi(MainWindow)
    MainWindow.show()
    sys.exit(app.exec_())


